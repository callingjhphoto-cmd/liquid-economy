"""Generate Excel tracker from live database data."""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import desc
from sqlalchemy.orm import Session
from database.models import (
    Company, Brand, BrandPrice, Valuation, DataSource,
    Indicator, ArbitrageSignal
)

# Design system colours
NAVY = "1A1F36"
GOLD = "C9A96E"
BLUE = "2B6CB0"
GREEN = "38A169"
LIGHT = "E2E8F0"
OFF_WHITE = "F7F8FA"
WHITE = "FFFFFF"

header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
header_fill = PatternFill("solid", fgColor=NAVY)
body_font = Font(name="Arial", size=9.5, color="333333")
thin_border = Border(
    left=Side(style="thin", color=LIGHT),
    right=Side(style="thin", color=LIGHT),
    top=Side(style="thin", color=LIGHT),
    bottom=Side(style="thin", color=LIGHT),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=OFF_WHITE if alt else WHITE)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.fill = fill
        cell.alignment = left_wrap
        cell.border = thin_border


def generate_tracker_excel(db: Session) -> io.BytesIO:
    """Generate the full tracker workbook from live data."""
    wb = Workbook()

    # ── SHEET 1: Company Profiles ──
    ws1 = wb.active
    ws1.title = "Company Profiles"
    ws1.sheet_properties.tabColor = BLUE

    headers = ["Company", "Ticker", "Public/Private", "Market Cap Tier",
               "Market Cap (USD)", "HQ", "FY End", "Key Brands", "IR URL"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    companies = db.query(Company).order_by(Company.name).all()
    for i, co in enumerate(companies):
        r = i + 2
        latest_val = (
            db.query(Valuation)
            .filter(Valuation.company_id == co.id)
            .order_by(desc(Valuation.date))
            .first()
        )

        ws1.cell(row=r, column=1, value=co.name)
        ws1.cell(row=r, column=2, value=co.ticker or "Private")
        ws1.cell(row=r, column=3, value="Public" if co.is_public else "Private")
        ws1.cell(row=r, column=4, value=co.market_cap_tier or "N/A")
        mcap = latest_val.market_cap if latest_val else None
        ws1.cell(row=r, column=5, value=mcap if mcap else "N/A")
        if isinstance(mcap, (int, float)):
            ws1.cell(row=r, column=5).number_format = '$#,##0'
        ws1.cell(row=r, column=6, value=co.hq_country)
        ws1.cell(row=r, column=7, value=co.fiscal_year_end)
        ws1.cell(row=r, column=8, value=co.key_brands)
        ws1.cell(row=r, column=9, value=co.ir_url)
        style_row(ws1, r, len(headers), alt=(i % 2 == 1))

    for c, w in enumerate([24, 16, 14, 14, 18, 16, 14, 50, 40], 1):
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.freeze_panes = "A2"

    # ── SHEET 2: Valuation Tracker ──
    ws2 = wb.create_sheet("Valuation Tracker")
    ws2.sheet_properties.tabColor = GREEN

    val_headers = ["Company", "Ticker", "Date", "Share Price", "Currency",
                   "EPS (TTM)", "P/E (TTM)", "P/E (Fwd)", "Div Yield", "Market Cap"]
    for c, h in enumerate(val_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(val_headers))

    public_cos = db.query(Company).filter(Company.is_public == True).order_by(Company.name).all()
    r = 2
    for co in public_cos:
        vals = (
            db.query(Valuation)
            .filter(Valuation.company_id == co.id)
            .order_by(desc(Valuation.date))
            .limit(12)  # Last 12 data points
            .all()
        )
        for v in vals:
            ws2.cell(row=r, column=1, value=co.name)
            ws2.cell(row=r, column=2, value=co.ticker)
            ws2.cell(row=r, column=3, value=v.date.isoformat())
            ws2.cell(row=r, column=4, value=v.share_price)
            ws2.cell(row=r, column=5, value=v.currency)
            ws2.cell(row=r, column=6, value=v.eps_ttm)
            ws2.cell(row=r, column=7, value=v.pe_ttm)
            if v.pe_ttm:
                ws2.cell(row=r, column=7).number_format = '0.0"x"'
            ws2.cell(row=r, column=8, value=v.pe_forward)
            if v.pe_forward:
                ws2.cell(row=r, column=8).number_format = '0.0"x"'
            ws2.cell(row=r, column=9, value=v.dividend_yield)
            if v.dividend_yield:
                ws2.cell(row=r, column=9).number_format = '0.0%'
            ws2.cell(row=r, column=10, value=v.market_cap)
            if v.market_cap:
                ws2.cell(row=r, column=10).number_format = '$#,##0'
            style_row(ws2, r, len(val_headers), alt=(r % 2 == 0))
            r += 1

    ws2.freeze_panes = "A2"

    # ── SHEET 3: Brand RRP Pricing ──
    ws3 = wb.create_sheet("Brand RRP Pricing")
    ws3.sheet_properties.tabColor = GOLD

    price_headers = ["Company", "Brand", "Expression", "Category", "Size",
                     "USA ($)", "UK (£)", "Europe (€)", "Middle East ($)",
                     "Price Diff", "Premium Index", "As Of"]
    for c, h in enumerate(price_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(price_headers))

    brands = (
        db.query(Brand, Company.name)
        .join(Company)
        .order_by(Company.name, Brand.brand_name)
        .all()
    )

    r = 2
    for brand, company_name in brands:
        latest_price = (
            db.query(BrandPrice)
            .filter(BrandPrice.brand_id == brand.id)
            .order_by(desc(BrandPrice.effective_date))
            .first()
        )
        if not latest_price:
            continue

        ws3.cell(row=r, column=1, value=company_name)
        ws3.cell(row=r, column=2, value=brand.brand_name)
        ws3.cell(row=r, column=3, value=brand.expression)
        ws3.cell(row=r, column=4, value=brand.category)
        ws3.cell(row=r, column=5, value=brand.size)
        ws3.cell(row=r, column=6, value=latest_price.price_usa)
        ws3.cell(row=r, column=7, value=latest_price.price_uk)
        ws3.cell(row=r, column=8, value=latest_price.price_eu)
        ws3.cell(row=r, column=9, value=latest_price.price_me)

        # Formulas for differential and premium index
        ws3.cell(row=r, column=10, value=f'=MAX(F{r}:I{r})-MIN(F{r}:I{r})')
        ws3.cell(row=r, column=10).number_format = '$#,##0'
        ws3.cell(row=r, column=11, value=f'=IF(MIN(F{r}:I{r})>0,MAX(F{r}:I{r})/MIN(F{r}:I{r})-1,"")')
        ws3.cell(row=r, column=11).number_format = '0.0%'
        ws3.cell(row=r, column=12, value=latest_price.effective_date.isoformat())

        style_row(ws3, r, len(price_headers), alt=(r % 2 == 0))
        for pc in [6, 7, 8, 9]:
            ws3.cell(row=r, column=pc).number_format = '#,##0'
            ws3.cell(row=r, column=pc).alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    ws3.freeze_panes = "A2"

    # ── SHEET 4: Arbitrage Signals ──
    ws4 = wb.create_sheet("Arbitrage Signals")
    ws4.sheet_properties.tabColor = GREEN

    arb_headers = ["Type", "Company/Brand", "Signal", "Severity", "Last Updated"]
    for c, h in enumerate(arb_headers, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(arb_headers))

    signals = db.query(ArbitrageSignal).filter(ArbitrageSignal.is_active == True).all()
    for i, s in enumerate(signals):
        r = i + 2
        ws4.cell(row=r, column=1, value=s.signal_type)
        ws4.cell(row=r, column=2, value=s.company_or_brand)
        ws4.cell(row=r, column=3, value=s.signal_description)
        ws4.cell(row=r, column=4, value=s.severity)
        ws4.cell(row=r, column=5, value=s.last_updated.isoformat() if s.last_updated else "")

        # Color-code severity
        sev_colors = {"critical": "C53030", "high": "DD6B20", "medium": GOLD, "low": "718096"}
        ws4.cell(row=r, column=4).font = Font(
            name="Arial", bold=True,
            color=sev_colors.get(s.severity, "333333"), size=9.5
        )
        style_row(ws4, r, len(arb_headers), alt=(i % 2 == 1))

    ws4.freeze_panes = "A2"

    # ── SHEET 5: Data Sources ──
    ws5 = wb.create_sheet("Data Sources")
    ws5.sheet_properties.tabColor = BLUE

    src_headers = ["Source", "Category", "URL", "Frequency", "Automation Tier",
                   "Last Collected", "Status"]
    for c, h in enumerate(src_headers, 1):
        ws5.cell(row=1, column=c, value=h)
    style_header(ws5, 1, len(src_headers))

    sources = db.query(DataSource).order_by(DataSource.automation_tier, DataSource.name).all()
    for i, s in enumerate(sources):
        r = i + 2
        ws5.cell(row=r, column=1, value=s.name)
        ws5.cell(row=r, column=2, value=s.category)
        ws5.cell(row=r, column=3, value=s.url)
        ws5.cell(row=r, column=4, value=s.update_frequency)
        ws5.cell(row=r, column=5, value=f"Tier {s.automation_tier}")
        ws5.cell(row=r, column=6, value=s.last_collected.isoformat() if s.last_collected else "Not yet")
        ws5.cell(row=r, column=7, value=s.status)
        style_row(ws5, r, len(src_headers), alt=(i % 2 == 1))

    ws5.freeze_panes = "A2"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
